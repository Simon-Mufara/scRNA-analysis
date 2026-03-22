"""
Export utilities for properly formatting and exporting analysis results.
Ensures CSV, Excel, and PDF exports are properly formatted with separated columns.
"""

import pandas as pd
import numpy as np
import io as _io
from typing import Optional
from fpdf import FPDF


def prepare_df_for_export(df: pd.DataFrame, numeric_cols: list = None,
                         precision: int = 4, scientific: bool = False) -> pd.DataFrame:
    """
    Prepare dataframe for export with proper numeric formatting.

    Args:
        df: DataFrame to prepare
        numeric_cols: List of numeric column names to format
        precision: Decimal places for rounding
        scientific: Use scientific notation for small p-values

    Returns:
        DataFrame with properly formatted values suitable for export
    """
    export_df = df.copy()

    if numeric_cols is None:
        numeric_cols = export_df.select_dtypes(include=[np.number]).columns.tolist()

    for col in numeric_cols:
        if col not in export_df.columns:
            continue

        if 'p' in col.lower() or 'pval' in col.lower():
            # P-values: use scientific notation
            export_df[col] = export_df[col].apply(
                lambda x: f"{x:.2e}" if isinstance(x, (int, float)) and x < 0.0001 else
                         f"{x:.4f}" if isinstance(x, (int, float)) else str(x)
            )
        else:
            # Other numeric: standard formatting
            export_df[col] = export_df[col].apply(
                lambda x: f"{x:.{precision}f}" if isinstance(x, (int, float)) else str(x)
            )

    return export_df


def export_to_excel(df: pd.DataFrame, sheet_name: str = "Results",
                   title: str = None, numeric_cols: list = None) -> bytes:
    """
    Export DataFrame to proper Excel format with formatting.

    Args:
        df: DataFrame to export
        sheet_name: Name for the Excel sheet
        title: Optional title row
        numeric_cols: Columns to format as numbers

    Returns:
        Excel file bytes
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Prepare data for export
    export_df = prepare_df_for_export(df, numeric_cols=numeric_cols)

    xlsx_buf = _io.BytesIO()
    with pd.ExcelWriter(xlsx_buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name=sheet_name)

        ws = writer.sheets[sheet_name]

        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_num, col_title in enumerate(export_df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Set column widths and center align
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col_num, col_title in enumerate(export_df.columns, 1):
            col_letter = ws.cell(row=1, column=col_num).column_letter
            max_len = max(
                len(str(col_title)),
                max((len(str(export_df.iloc[i, col_num-1])) for i in range(len(export_df))), default=0)
            )
            adjusted_width = min(max_len + 3, 50)
            ws.column_dimensions[col_letter].width = adjusted_width

            # Center align all cells
            for row_num in range(2, len(export_df) + 2):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

        # Freeze header row
        ws.freeze_panes = "A2"

    return xlsx_buf.getvalue()


def export_to_csv(df: pd.DataFrame, numeric_cols: list = None) -> bytes:
    """
    Export DataFrame to proper CSV format.
    Ensures columns are properly separated.

    Args:
        df: DataFrame to export
        numeric_cols: Columns to format

    Returns:
        CSV file bytes (UTF-8 encoded)
    """
    # Prepare data
    export_df = prepare_df_for_export(df, numeric_cols=numeric_cols)

    # Convert to CSV with proper delimiter and encoding
    csv_bytes = export_df.to_csv(
        index=False,
        sep=',',  # Standard comma separator
        quoting=1,  # Quote minimal fields
        doublequote=True,
        lineterminator='\n',
        encoding='utf-8'
    ).encode('utf-8')

    return csv_bytes


class ReportPDF(FPDF):
    """Enhanced PDF class for scientific reports."""

    def __init__(self, title: str = "Analysis Report"):
        super().__init__()
        self.title = title
        self.page_number = 0

    def header(self):
        """Add header to each page."""
        if self.page_no() > 1:
            self.set_font("Arial", "", 8)
            self.set_text_color(128, 128, 128)
            self.cell(0, 10, f"{self.title} — Page {self.page_no()}", 0, 1, "R")
            self.ln(2)

    def footer(self):
        """Add footer to each page."""
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        self.set_text_color(128, 128, 128)
        self.cell(0, 10, f"Page {self.page_no()}", 0, 0, "C")

    def add_section_title(self, text: str, level: int = 1):
        """Add a formatted section title."""
        if level == 1:
            self.set_font("Arial", "B", 14)
            self.set_text_color(0, 0, 0)
            self.ln(5)
        elif level == 2:
            self.set_font("Arial", "B", 12)
            self.set_text_color(0, 0, 0)
            self.ln(3)
        else:
            self.set_font("Arial", "B", 10)
            self.set_text_color(50, 50, 50)
            self.ln(2)

        self.cell(0, 8, text, 0, 1)
        self.ln(2)

    def add_metric(self, label: str, value: str):
        """Add a metric in two-column layout."""
        self.set_font("Arial", "B", 10)
        self.set_text_color(0, 0, 0)

        # Label column (40% width)
        self.set_x(self.l_margin)
        self.cell(self.w * 0.4 - self.l_margin, 6, label, 0)

        # Value column (60% width)
        self.set_font("Arial", "", 10)
        self.set_text_color(0, 0, 100)
        self.cell(self.w * 0.6, 6, str(value), 0, 1)

    def add_table(self, df: pd.DataFrame, column_widths: list = None):
        """Add a formatted table to the PDF."""
        self.set_font("Arial", "B", 9)
        self.set_text_color(255, 255, 255)
        self.set_fill_color(66, 114, 219)

        # Calculate column widths
        if column_widths is None:
            col_width = (self.w - self.l_margin - self.r_margin) / len(df.columns)
            column_widths = [col_width] * len(df.columns)

        # Header
        for col_num, col_title in enumerate(df.columns):
            self.cell(column_widths[col_num], 7, str(col_title)[:20], 1, 0, "C", True)
        self.ln()

        # Rows
        self.set_font("Arial", "", 8)
        self.set_text_color(0, 0, 0)

        for idx, row in df.iterrows():
            for col_num, col_title in enumerate(df.columns):
                value = str(row[col_title])[:30]  # Truncate long values
                self.cell(column_widths[col_num], 6, value, 1, 0, "C")
            self.ln()

        self.ln(3)

    def add_text(self, text: str, size: int = 10, bold: bool = False):
        """Add body text with automatic wrapping."""
        self.set_font("Arial", "B" if bold else "", size)
        self.set_text_color(0, 0, 0)
        self.multi_cell(0, 5, text)
        self.ln(2)


def create_analysis_report_pdf(adata, title: str = "scRNA-seq Analysis Report",
                              include_sections: dict = None) -> bytes:
    """
    Create a professional analysis report PDF.

    Args:
        adata: AnnData object with analysis results
        title: Report title
        include_sections: Dict specifying which sections to include

    Returns:
        PDF file bytes
    """
    if include_sections is None:
        include_sections = {
            "summary": True,
            "qc": True,
            "clustering": True,
            "annotation": True,
        }

    pdf = ReportPDF(title=title)
    pdf.add_page()

    # Title page
    pdf.set_font("Arial", "B", 20)
    pdf.set_text_color(0, 0, 100)
    pdf.ln(20)
    pdf.cell(0, 15, title, 0, 1, "C")

    pdf.set_font("Arial", "I", 11)
    pdf.set_text_color(100, 100, 100)
    from datetime import datetime
    pdf.cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", 0, 1, "C")

    pdf.ln(10)

    # Summary section
    if include_sections.get("summary"):
        pdf.add_section_title("Dataset Summary")
        pdf.add_metric("Total Cells", f"{adata.n_obs:,}")
        pdf.add_metric("Total Genes", f"{adata.n_vars:,}")

        if "leiden" in adata.obs.columns:
            n_clusters = adata.obs["leiden"].nunique()
            pdf.add_metric("Clusters Identified", f"{n_clusters}")

        if "cell_type" in adata.obs.columns:
            n_types = adata.obs["cell_type"].nunique()
            pdf.add_metric("Cell Types", f"{n_types}")

    # QC section
    if include_sections.get("qc"):
        pdf.add_section_title("Quality Control Metrics", level=2)

        if "n_genes_by_counts" in adata.obs.columns:
            median_genes = adata.obs["n_genes_by_counts"].median()
            pdf.add_metric("Median Genes per Cell", f"{median_genes:,.0f}")

        if "total_counts" in adata.obs.columns:
            median_counts = adata.obs["total_counts"].median()
            pdf.add_metric("Median UMI per Cell", f"{median_counts:,.0f}")

        if "pct_counts_mt" in adata.obs.columns:
            median_mito = adata.obs["pct_counts_mt"].median()
            pdf.add_metric("Median Mitochondrial %", f"{median_mito:.2f}%")

    # Clustering section
    if include_sections.get("clustering") and "leiden" in adata.obs.columns:
        pdf.add_section_title("Clustering Results", level=2)

        cluster_counts = adata.obs["leiden"].value_counts().sort_index().head(10)
        cluster_data = pd.DataFrame({
            "Cluster": cluster_counts.index.astype(str),
            "# Cells": cluster_counts.values.astype(str),
            "% Cells": (100 * cluster_counts.values / adata.n_obs).round(1).astype(str)
        })

        pdf.add_text(f"Identified {len(adata.obs['leiden'].unique())} clusters (showing top 10)")
        pdf.add_table(cluster_data, column_widths=[20, 30, 30])

    # Annotation section
    if include_sections.get("annotation") and "cell_type" in adata.obs.columns:
        pdf.add_section_title("Cell Type Composition", level=2)

        cell_type_counts = adata.obs["cell_type"].value_counts().head(10)
        celltype_data = pd.DataFrame({
            "Cell Type": cell_type_counts.index[:10],
            "Count": cell_type_counts.values[:10].astype(str),
            "Percentage": (100 * cell_type_counts.values[:10] / adata.n_obs).round(1).astype(str) + "%"
        })

        pdf.add_table(celltype_data, column_widths=[50, 25, 30])

    return pdf.output(dest='S').encode('latin-1')
